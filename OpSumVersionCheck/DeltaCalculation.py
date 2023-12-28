import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string


def delta_calculation(excel_file_path, start_point, sheet_name_current, sheet_name_previous):
    sheet_name_delta = sheet_name_current + " delta"

    header_index = start_point[0] - 1
    column_start_index = column_index_from_string(start_point[1]) - 1

    # Read the Excel file into a Pandas DataFrame
    df_current = pd.read_excel(excel_file_path, sheet_name=sheet_name_current, header=header_index)
    df_previous = pd.read_excel(excel_file_path, sheet_name=sheet_name_previous, header=header_index)

    # # Get the header names
    header_row_current = df_current.columns
    header_row_previous = df_previous.columns

    delta_data = {}

    current_column_index = 0
    for header_name_current in header_row_current:

        if current_column_index < column_start_index:
            current_column_index = current_column_index + 1

            # columns before the actual data, just copy into the new form
            delta_data[header_name_current] = df_current[header_name_current]

            continue
        current_column_index = current_column_index + 1

        previous_column_index = -1
        for header_name_previous in header_row_previous:
            previous_column_index = previous_column_index + 1

            if header_name_current == header_name_previous:

                # found the match
                column_values_current = df_current[header_name_current]

                result_list = []
                for i in range(len(column_values_current)):
                    row_index = i + start_point[0] + 1
                    column_letter = get_column_letter(current_column_index)
                    previous_column_letter = get_column_letter(previous_column_index)

                    result_list.append(
                        f"='{sheet_name_current}'!{column_letter}{row_index}-'{sheet_name_previous}'!{previous_column_letter}{row_index}")

                delta_data[header_name_current] = result_list

        # Add empty column if the header doesn't match any header in sheet 2
        if header_name_current not in delta_data:
            empty_array = []
            count = len(df_current[header_name_current])
            for i in range(count):
                empty_array.append(None)
            delta_data[header_name_current] = empty_array

    # Write the data into the Delta sheet
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        result_df = pd.DataFrame(delta_data)
        result_df.to_excel(writer, sheet_name=sheet_name_delta, index=False, startrow=header_index)
