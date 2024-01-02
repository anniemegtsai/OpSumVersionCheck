import os

from openpyxl import load_workbook
from pathlib import Path
import shutil
from typing import List
from Utils.modules import Sheet
import pandas as pd


def create_values_only_excel_file(input_file, output_file):
    wb = load_workbook(input_file, data_only=True)
    wb.save(output_file)


def is_file_exist(file_path):
    file = Path(file_path)
    if file.is_file():
        return True
    else:
        return False


def delete_file(file_path):
    file = Path(file_path)

    if file.is_file():
        file.unlink()


def create_file_copy(input_file, output_file):
    shutil.copyfile(input_file, output_file)


def create_directory(directory_path):
    os.mkdir(directory_path)


def delete_directory(directory_path):
    file = Path(directory_path)

    if file.is_dir():
        shutil.rmtree(directory_path)


def get_file_name_from_file_path(file_path):
    return Path(file_path).name


def get_file_name_without_extension_from_file_path(file_path):
    return Path(file_path).stem


def quit_excel():
    os.system('taskkill /T /IM EXCEL.exe')


def create_sheet(excel_file_path, sheet_name):
    wb = load_workbook(excel_file_path)
    wb.create_sheet(sheet_name)
    wb.save(excel_file_path)


def write_sheets_to_excel(excel_file_path: str, sheet_list: List[Sheet]):
    with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
        for sheet in sheet_list:
            sheet.data_frame.to_excel(excel_writer=writer, sheet_name=sheet.name, index=False,
                                      startrow=sheet.start_point[0] - 1)


def load_all_sheets(excel_file_path):
    xlsx = pd.ExcelFile(excel_file_path)

    sheet_list = []
    for sheet_name in xlsx.sheet_names:
        data_frame = xlsx.parse(sheet_name)
        sheet_list.append(Sheet(name=sheet_name, data_frame=data_frame))

    return sheet_list


def replace_extension(file_name, extension):
    return file_name.rsplit('.', 1)[0] + extension
