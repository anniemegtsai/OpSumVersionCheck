import os
from openpyxl import load_workbook

# 指定目錄路徑
directory_path = r'C:\Users\annie\Desktop\Python\Test\Combine'

# 創建一個新的 Excel 檔案
output_workbook = None

# 迴圈遍歷目錄中的每個 Excel 檔案
for filename in os.listdir(directory_path):
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # 載入每個 Excel 檔案的 workbook
        current_workbook = load_workbook(os.path.join(directory_path, filename), read_only=False)

        # 如果是第一個檔案，則複製整個 workbook
        if output_workbook is None:
            output_workbook = current_workbook
        else:
            # 否則將每個工作表複製到新的檔案中
            for sheet_name in current_workbook.sheetnames:
                # 複製工作表
                current_sheet = current_workbook[sheet_name]
                new_sheet = output_workbook.create_sheet(title=f"{filename}_{sheet_name}")

                # 複製工作表的內容
                for row in current_sheet.iter_rows(values_only=True):
                    new_sheet.append(row)

# 保存合併後的 Excel 檔案
output_path = r'C:\Users\annie\Desktop\Python\Test\merged_workbook.xlsx'
output_workbook.save(output_path)

print(f'合併完成，結果保存到 {output_path}')
