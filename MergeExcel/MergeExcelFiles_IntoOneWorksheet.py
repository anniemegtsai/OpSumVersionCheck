import os
import openpyxl

# 設定目錄路徑
directory_path = r'C:\Users\annie\Desktop\Python\Test\Combine'  # 替換為你的目錄路徑

# 創建新的工作簿和工作表
merged_wb = openpyxl.Workbook()
merged_ws = merged_wb.active

# 迭代目錄中的所有檔案
for filename in os.listdir(directory_path):
    if filename.endswith('.xlsx'):  # 確認是Excel檔案
        file_path = os.path.join(directory_path, filename)

        # 開啟Excel檔案
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 獲取第一個資料表（你也可以根據需要指定其他資料表）
        ws = wb.active

        # 迭代並將每一行的資料複製到新的資料表
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            merged_ws.append(row)


os.chdir(r'C:\Users\annie\Desktop\Python\Test')
# 保存合併後的Excel檔案
merged_wb.save('merged_data.xlsx')